# from openpyxl import Workbook
# wb = Workbook()
#
# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
#
# # Save the file
# wb.save("sample.xlsx")


import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# create empty data structures
products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

# how many rows
# print(product_list.max_row)

# loop through entire sheet
for product_row in range(2, product_list.max_row + 1):
    # get values for each column
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value

    # add new value in the file
    inventory_price = product_list.cell(product_row, 5)

    # calculation number of products per supplier
    """using one line if else"""
    # products_per_supplier[supplier_name] = products_per_supplier[
    #                                            supplier_name] + 1 if supplier_name in products_per_supplier else 1

    # calculation number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_product = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_product + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # add value for total inventory price
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)


inv_file.save("inventory_with_total_value.xlsx")

